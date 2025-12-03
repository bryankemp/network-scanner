"""
Report generation for network scan results.
Extracted and adapted from network_scan/mapper.py
"""

import os
import subprocess
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generate_html_report(hosts: List[Dict[str, Any]], output_file: str) -> str:
    """
    Generate HTML report with links to all services.

    Args:
        hosts: List of host dictionaries from parser
        output_file: Path to output HTML file

    Returns:
        Path to generated HTML file
    """
    html = (
        """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Network Map Report</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #2c3e50;
            border-bottom: 3px solid #3498db;
            padding-bottom: 10px;
        }
        h2 {
            color: #34495e;
            margin-top: 30px;
        }
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }
        .stat-card {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
        }
        .stat-card h3 {
            margin: 0;
            font-size: 2em;
        }
        .stat-card p {
            margin: 5px 0 0 0;
            opacity: 0.9;
        }
        .host-card {
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
        }
        .host-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 15px;
        }
        .host-title {
            font-size: 1.3em;
            font-weight: bold;
            color: #2c3e50;
        }
        .vm-badge {
            background-color: #e74c3c;
            color: white;
            padding: 5px 10px;
            border-radius: 4px;
            font-size: 0.9em;
            font-weight: bold;
        }
        .host-info {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 10px;
            margin-bottom: 15px;
        }
        .info-item {
            padding: 8px;
            background-color: white;
            border-radius: 4px;
        }
        .info-label {
            font-weight: bold;
            color: #7f8c8d;
            font-size: 0.9em;
        }
        .services-table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 10px;
        }
        .services-table th {
            background-color: #34495e;
            color: white;
            padding: 10px;
            text-align: left;
        }
        .services-table td {
            padding: 10px;
            border-bottom: 1px solid #dee2e6;
        }
        .services-table tr:hover {
            background-color: #ecf0f1;
        }
        .link {
            color: #3498db;
            text-decoration: none;
            font-weight: bold;
        }
        .link:hover {
            text-decoration: underline;
        }
        .no-services {
            color: #95a5a6;
            font-style: italic;
            padding: 10px;
        }
        .footer {
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #dee2e6;
            text-align: center;
            color: #7f8c8d;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>🌐 Network Map Report</h1>
        <p><strong>Generated:</strong> """
        + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        + """</p>
"""
    )

    # Calculate statistics
    total_hosts = len(hosts)
    total_vms = sum(1 for h in hosts if h.get("is_vm"))
    total_services = sum(len(h.get("ports", [])) for h in hosts)

    html += f"""
        <div class="stats">
            <div class="stat-card">
                <h3>{total_hosts}</h3>
                <p>Total Devices</p>
            </div>
            <div class="stat-card">
                <h3>{total_vms}</h3>
                <p>Virtual Machines</p>
            </div>
            <div class="stat-card">
                <h3>{total_services}</h3>
                <p>Services Found</p>
            </div>
        </div>
        
        <h2>Discovered Devices</h2>
"""

    # Sort hosts by IP
    hosts_sorted = sorted(hosts, key=lambda x: tuple(map(int, x.get("ip", "0.0.0.0").split("."))))

    for host in hosts_sorted:
        hostname_display = host.get("hostname") or "Unknown"
        vm_badge = (
            f'<span class="vm-badge">🖥️ {host.get("vm_type", "VM")}</span>'
            if host.get("is_vm")
            else ""
        )

        html += f"""
        <div class="host-card">
            <div class="host-header">
                <div class="host-title">{hostname_display} ({host.get('ip', 'N/A')})</div>
                {vm_badge}
            </div>
            <div class="host-info">
                <div class="info-item">
                    <div class="info-label">IP Address</div>
                    <div>{host.get('ip', 'N/A')}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">MAC Address</div>
                    <div>{host.get('mac') or 'N/A'}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">Vendor</div>
                    <div>{host.get('vendor') or 'N/A'}</div>
                </div>
"""

        if host.get("os"):
            html += f"""
                <div class="info-item">
                    <div class="info-label">Operating System</div>
                    <div>{host['os']}</div>
                </div>
"""

        html += """
            </div>
"""

        if host.get("ports"):
            html += """
            <h3>Services</h3>
            <table class="services-table">
                <thead>
                    <tr>
                        <th>Port</th>
                        <th>Protocol</th>
                        <th>Service</th>
                        <th>Details</th>
                        <th>Access Link</th>
                    </tr>
                </thead>
                <tbody>
"""

            for port in host["ports"]:
                details = []
                if port.get("product"):
                    details.append(port["product"])
                if port.get("version"):
                    details.append(port["version"])
                if port.get("extrainfo"):
                    details.append(port["extrainfo"])
                details_str = " ".join(details) if details else "N/A"

                # Generate links for common services
                link = ""
                service = port.get("service", "")
                if service in ["http", "https"] or "http" in service:
                    protocol = "https" if "ssl" in service or service == "https" else "http"
                    port_num = port.get("port")
                    url = f"{protocol}://{host.get('ip')}:{port_num}"
                    link = f'<a href="{url}" class="link" target="_blank">Open ↗</a>'
                elif service == "ssh":
                    link = f'<code>ssh {host.get("ip")} -p {port.get("port")}</code>'

                html += f"""
                    <tr>
                        <td>{port.get('port')}</td>
                        <td>{port.get('protocol', '').upper()}</td>
                        <td><strong>{service}</strong></td>
                        <td>{details_str}</td>
                        <td>{link}</td>
                    </tr>
"""

            html += """
                </tbody>
            </table>
"""
        else:
            html += '<div class="no-services">No services detected</div>'

        html += """
        </div>
"""

    html += """
        <div class="footer">
            <p>Network Map Report | Generated by Network Scanner</p>
        </div>
    </div>
</body>
</html>
"""

    os.makedirs(
        os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True
    )
    with open(output_file, "w") as f:
        f.write(html)

    return output_file


def generate_xlsx_report(hosts: List[Dict[str, Any]], output_file: str) -> str:
    """
    Generate Excel spreadsheet with network data.

    Args:
        hosts: List of host dictionaries from parser
        output_file: Path to output XLSX file

    Returns:
        Path to generated XLSX file
    """
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    ws_summary["A1"] = "Network Map Summary"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A3"] = "Generated:"
    ws_summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws_summary["A5"] = "Total Devices:"
    ws_summary["B5"] = len(hosts)
    ws_summary["A6"] = "Virtual Machines:"
    ws_summary["B6"] = sum(1 for h in hosts if h.get("is_vm"))
    ws_summary["A7"] = "Total Services:"
    ws_summary["B7"] = sum(len(h.get("ports", [])) for h in hosts)

    # Devices sheet
    ws_devices = wb.create_sheet("Devices")
    headers = [
        "IP Address",
        "Hostname",
        "MAC Address",
        "Vendor",
        "Is VM",
        "VM Type",
        "Open Ports",
        "OS",
    ]
    ws_devices.append(headers)

    for cell in ws_devices[1]:
        cell.fill = header_fill
        cell.font = header_font

    for host in sorted(hosts, key=lambda x: tuple(map(int, x.get("ip", "0.0.0.0").split(".")))):
        ws_devices.append(
            [
                host.get("ip", ""),
                host.get("hostname", ""),
                host.get("mac", ""),
                host.get("vendor", ""),
                "Yes" if host.get("is_vm") else "No",
                host.get("vm_type", ""),
                len(host.get("ports", [])),
                host.get("os", ""),
            ]
        )

    # Auto-adjust column widths
    for column in ws_devices.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (AttributeError, TypeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_devices.column_dimensions[column_letter].width = adjusted_width

    # Services sheet
    ws_services = wb.create_sheet("Services")
    headers = [
        "IP Address",
        "Hostname",
        "Port",
        "Protocol",
        "Service",
        "Product",
        "Version",
        "Extra Info",
    ]
    ws_services.append(headers)

    for cell in ws_services[1]:
        cell.fill = header_fill
        cell.font = header_font

    for host in sorted(hosts, key=lambda x: tuple(map(int, x.get("ip", "0.0.0.0").split(".")))):
        for port in host.get("ports", []):
            ws_services.append(
                [
                    host.get("ip", ""),
                    host.get("hostname", ""),
                    port.get("port", ""),
                    port.get("protocol", "").upper(),
                    port.get("service", ""),
                    port.get("product", ""),
                    port.get("version", ""),
                    port.get("extrainfo", ""),
                ]
            )

    # Auto-adjust column widths
    for column in ws_services.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (AttributeError, TypeError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_services.column_dimensions[column_letter].width = adjusted_width

    # Virtual Machines sheet
    ws_vms = wb.create_sheet("Virtual Machines")
    headers = ["IP Address", "Hostname", "VM Type", "Vendor", "MAC Address", "Services"]
    ws_vms.append(headers)

    for cell in ws_vms[1]:
        cell.fill = header_fill
        cell.font = header_font

    for host in sorted(
        [h for h in hosts if h.get("is_vm")],
        key=lambda x: tuple(map(int, x.get("ip", "0.0.0.0").split("."))),
    ):
        services_list = ", ".join(
            [f"{p.get('port')}/{p.get('service')}" for p in host.get("ports", [])]
        )
        ws_vms.append(
            [
                host.get("ip", ""),
                host.get("hostname", ""),
                host.get("vm_type", ""),
                host.get("vendor", ""),
                host.get("mac", ""),
                services_list,
            ]
        )

    # Auto-adjust column widths
    for column in ws_vms.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (AttributeError, TypeError):
                pass
        adjusted_width = min(max_length + 2, 60)
        ws_vms.column_dimensions[column_letter].width = adjusted_width

    os.makedirs(
        os.path.dirname(output_file) if os.path.dirname(output_file) else ".", exist_ok=True
    )
    wb.save(output_file)

    return output_file


def generate_graphviz_diagram(
    hosts: List[Dict[str, Any]], output_base: str
) -> tuple[str, str, str]:
    """
    Generate network topology diagram using Graphviz.

    Args:
        hosts: List of host dictionaries from parser
        output_base: Base path for output files (without extension)

    Returns:
        Tuple of (dot_file_path, png_file_path, svg_file_path)
    """
    dot_content = """digraph NetworkMap {
    rankdir=TB;
    node [shape=box, style=filled, fillcolor=lightblue];
    edge [color=gray];
    
    // Gateway/Router
    gateway [label="Gateway\\n192.168.1.1", fillcolor="#e74c3c", fontcolor=white, fontsize=12, shape=box3d];
    
"""

    # Group hosts by type
    vms = [h for h in hosts if h.get("is_vm")]
    physical = [h for h in hosts if not h.get("is_vm")]

    # Add physical hosts
    for host in physical:
        if host.get("ip") == "192.168.1.1":
            continue  # Skip gateway

        label = f"{host.get('hostname') or 'Unknown'}\\n{host.get('ip')}"
        if host.get("ports"):
            top_services = [p.get("service") for p in host.get("ports", [])[:3]]
            label += "\\n" + ", ".join(top_services)

        color = "#3498db" if host.get("ports") else "#95a5a6"
        ip_safe = host.get("ip", "").replace(".", "_")
        dot_content += f'    host_{ip_safe} [label="{label}", fillcolor="{color}"];\n'
        dot_content += f"    gateway -> host_{ip_safe};\n"

    # Add VMs
    if vms:
        dot_content += "\n    // Virtual Machines\n"
        for host in vms:
            label = f"VM: {host.get('hostname') or 'Unknown'}\\n{host.get('ip')}\\n({host.get('vm_type')})"
            if host.get("ports"):
                top_services = [p.get("service") for p in host.get("ports", [])[:3]]
                label += "\\n" + ", ".join(top_services)

            ip_safe = host.get("ip", "").replace(".", "_")
            dot_content += (
                f'    vm_{ip_safe} [label="{label}", fillcolor="#9b59b6", fontcolor=white];\n'
            )
            dot_content += f"    gateway -> vm_{ip_safe};\n"

    dot_content += "}\n"

    # Write DOT file
    dot_file = f"{output_base}.dot"
    os.makedirs(os.path.dirname(dot_file) if os.path.dirname(dot_file) else ".", exist_ok=True)
    with open(dot_file, "w") as f:
        f.write(dot_content)

    png_file = f"{output_base}.png"
    svg_file = f"{output_base}.svg"

    # Try to generate PNG and SVG using dot command
    try:
        subprocess.run(["dot", "-Tpng", dot_file, "-o", png_file], check=True)
    except (subprocess.CalledProcessError, FileNotFoundError):
        print("⚠ Graphviz 'dot' command not found or failed. Install with: brew install graphviz")
        png_file = None

    try:
        subprocess.run(["dot", "-Tsvg", dot_file, "-o", svg_file], check=True)
    except (subprocess.CalledProcessError, FileNotFoundError):
        svg_file = None

    return (dot_file, png_file, svg_file)
